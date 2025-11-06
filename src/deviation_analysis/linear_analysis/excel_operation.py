import os
from typing import Optional
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

# Layout constants
EXCAVATION_COL_START = 7
DUMP_COL_START = 20
CELL_WIDTH_PX = 70
CELL_HEIGHT_PX = 50
CHART_SIZE_COLS = 10
CHART_SIZE_ROWS = 10
TITLE_ROW = 2
DATA_START_ROW = 3
CHART_OFFSET_ROWS = 2


def _add_summary_section(
    ws,
    df: pd.DataFrame,
    img_path: str,
    col_start: int,
    title: str
) -> None:
    """Add a summary section with table and chart to worksheet."""
    # Add title
    ws.cell(row=TITLE_ROW, column=col_start, value=title)

    # Add table
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True),
        start=DATA_START_ROW
    ):
        for c_idx, value in enumerate(row, start=col_start):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Add chart
    img = XLImage(img_path)
    img.width = CHART_SIZE_COLS * CELL_WIDTH_PX
    img.height = CHART_SIZE_ROWS * CELL_HEIGHT_PX
    chart_row = DATA_START_ROW + len(df) + CHART_OFFSET_ROWS
    ws.add_image(img, ws.cell(row=chart_row, column=col_start).coordinate)



def merge_and_save_dataframes(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    output_path: str
) -> str:
    """
    Merge two DataFrames and save to Excel.
    
    Args:
        df1: First DataFrame
        df2: Second DataFrame
        output_path: Output path (file or directory)
        
    Returns:
        Path to saved Excel file
        
    Raises:
        TypeError: If inputs are not DataFrames
        ValueError: If output_path is invalid
        IOError: If file save fails
    """
    # Validate inputs
    if not isinstance(df1, pd.DataFrame) or not isinstance(df2, pd.DataFrame):
        raise TypeError("Both df1 and df2 must be pandas DataFrames.")
    if not output_path or not isinstance(output_path, str):
        raise ValueError("output_path must be a valid string.")

    # Merge DataFrames
    merged_df = pd.concat([df1, df2], ignore_index=True)

    # Determine output filename
    if os.path.isdir(output_path):
        output_file = os.path.join(output_path, "merged_summary.xlsx")
    else:
        output_file = output_path if output_path.endswith(".xlsx") else f"{output_path}.xlsx"

    # Save to Excel
    try:
        merged_df.to_excel(output_file, index=False)
    except Exception as e:
        raise IOError(f"Failed to save Excel file: {e}")

    return output_file


def update_excel_with_tables_charts(
    input_excel: str,
    df_excavation: pd.DataFrame,
    img_path_excavation: str,
    df_dump: pd.DataFrame,
    img_path_dump: str
) -> str:
    """
    Update Excel workbook with summary tables and charts.
    
    Args:
        input_excel: Path to input Excel file
        df_excavation: Excavation summary DataFrame
        img_path_excavation: Path to excavation chart image
        df_dump: Dump summary DataFrame
        img_path_dump: Path to dump chart image
        
    Returns:
        Path to updated Excel file
    """
    # Load workbook
    wb = openpyxl.load_workbook(input_excel)
    ws = wb.active

    # Add excavation data
    _add_summary_section(
        ws,
        df_excavation,
        img_path_excavation,
        EXCAVATION_COL_START,
        "Excavation Summary"
    )

    # Add dump data
    _add_summary_section(
        ws,
        df_dump,
        img_path_dump,
        DUMP_COL_START,
        "Dump Summary"
    )

    # Save workbook
    wb.save(input_excel)
    return input_excel


